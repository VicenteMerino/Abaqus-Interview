import os

from django.db import transaction
from openpyxl import load_workbook

from styleguide_example.portfolio.models import Asset, Portfolio, PortfolioAsset, AssetPrice


class ETLService:
    """
    This class is responsible for extracting, transforming and loading data from
    excel files into the database.
    """

    def __init__(self, file):
        self.file = file
        self.workbook = load_workbook(file)
        self.weights_worksheet = self.workbook["weights"]
        self.prices_worksheet = self.workbook["Precios"]
        self.assets_dict = {}
        self.portfolios_list = []

    def _is_valid_path(self):
        return os.path.exists(self.file) and (
            self.file.endswith(".xlsx") or self.file.endswith(".xls")
        )

    def _extract_portfolios(self):
        portfolios = []
        for col in self.weights_worksheet.iter_cols(min_col=3):
            name = col[0].value
            if name is None:
                continue
            portfolios.append(Portfolio(name=name))
        portfolios_objs = Portfolio.objects.bulk_create(portfolios)
        self.portfolios_list.extend(portfolios_objs)
        for portfolio in portfolios_objs:
            print(f"Created portfolio {portfolio}")

    def _extract_assets(self):
        assets = []
        for row in self.weights_worksheet.iter_rows(min_row=2):
            asset_name = row[1].value
            if asset_name is None:
                continue
            assets.append(Asset(name=asset_name))
        assets_objs = Asset.objects.bulk_create(assets)
        for asset in assets_objs:
            print(f"Created asset {asset}")
            self.assets_dict[asset.name] = asset

    def _extract_weights(self):
        portfolio_assets = []
        for row in self.weights_worksheet.iter_rows(min_row=2):
            date = row[0].value
            if date is None:
                continue
            asset = self.assets_dict[row[1].value]
            for i, portfolio in enumerate(self.portfolios_list):
                weight = row[i + 2].value
                portfolio_assets.append(
                    PortfolioAsset(portfolio=portfolio, asset=asset, date=date, weight=weight)
                )
        portfolio_assets_objs = PortfolioAsset.objects.bulk_create(portfolio_assets)
        for portfolio_asset in portfolio_assets_objs:
            print(
                f"Created portfolio asset for {portfolio_asset} with weight {portfolio_asset.weight}"
            )

    def _extract_prices(self):
        column_names = [
            cell.value for cell in self.prices_worksheet[1][1:] if cell.value is not None
        ]
        asset_prices = []

        for row in self.prices_worksheet.iter_rows(min_row=2):
            date = row[0].value
            if date is None:
                continue
            for i, column_name in enumerate(column_names):
                asset_name = column_name
                price = row[i + 1].value
                asset_prices.append(
                    AssetPrice(asset=self.assets_dict[asset_name], date=date, value=price)
                )
        asset_prices_obj = AssetPrice.objects.bulk_create(asset_prices)
        for asset_price in asset_prices_obj:
            print(f"Created asset price for {asset_price} with value {asset_price.value}")

    @transaction.atomic
    def extract(self):
        if not self._is_valid_path():
            raise FileNotFoundError("File not found")

        self._extract_portfolios()
        self._extract_assets()
        self._extract_weights()
        self._extract_prices()
